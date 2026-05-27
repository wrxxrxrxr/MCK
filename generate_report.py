import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, 
    NamedStyle, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

#  РУССКИЕ МЕСЯЦА
RUSSIAN_MONTHS = {
    'January': 'Январь', 'February': 'Февраль', 'March': 'Март',
    'April': 'Апрель', 'May': 'Май', 'June': 'Июнь',
    'July': 'Июль', 'August': 'Август', 'September': 'Сентябрь',
    'October': 'Октябрь', 'November': 'Ноябрь', 'December': 'Декабрь'
}

def translate_month(month_name):
    """Переводит название месяца на русский"""
    if not month_name:
        return ''
    # Убираем возможный год из строки
    parts = month_name.split()
    month = parts[0]
    return RUSSIAN_MONTHS.get(month, month)

def format_date_range(date_from, date_to):
    """Форматирует диапазон дат для отображения"""
    try:
        d1 = datetime.fromisoformat(date_from)
        d2 = datetime.fromisoformat(date_to)
        return f"{d1.strftime('%d.%m.%Y')} — {d2.strftime('%d.%m.%Y')}"
    except:
        return f"{date_from} — {date_to}"

def format_number(value):
    """Форматирует число с пробелами как разделителями тысяч"""
    try:
        return f"{int(float(value)):,}".replace(",", " ")
    except:
        return str(value) if value else "0"

def format_currency(value):
    """Форматирует валюту BYN"""
    try:
        return f"{float(value):,.2f}".replace(",", " ").replace(".", ",") + " BYN"
    except:
        return "0,00 BYN"

def parse_date(val):
    """Безопасный парсинг даты"""
    if not val:
        return None
    try:
        if isinstance(val, str):
            return datetime.fromisoformat(val.replace("Z", ""))
        return val
    except:
        return None

def fmt_date(val):
    """Форматирует дату в формат ДД.ММ.ГГГГ"""
    dt = parse_date(val)
    if dt:
        return dt.strftime("%d.%m.%Y")
    return "—"

# СТИЛИ 
def get_styles():
    """Возвращает набор стилей для Excel"""
    styles = {}
    
    # Заголовок отчёта
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(bold=True, size=18, color="C0392B", name="Calibri")
    title_style.alignment = Alignment(horizontal="left", vertical="center")
    styles['title'] = title_style
    
    # Подзаголовок (период)
    period_style = NamedStyle(name="period_style")
    period_style.font = Font(size=11, color="666666", name="Calibri")
    period_style.alignment = Alignment(horizontal="left", vertical="center")
    styles['period'] = period_style
    
    # Заголовок секции
    section_style = NamedStyle(name="section_style")
    section_style.font = Font(bold=True, size=12, color="1A1A2E", name="Calibri")
    section_style.fill = PatternFill("solid", fgColor="F0F4F8")
    section_style.alignment = Alignment(horizontal="left", vertical="center")
    styles['section'] = section_style
    
    # Шапка таблицы
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    header_style.fill = PatternFill("solid", fgColor="C0392B")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = Border(
        left=Side(style="thin", color="A93226"),
        right=Side(style="thin", color="A93226"),
        top=Side(style="thin", color="A93226"),
        bottom=Side(style="thin", color="A93226")
    )
    styles['header'] = header_style
    
    # Ячейка данных (чётная строка)
    data_even_style = NamedStyle(name="data_even_style")
    data_even_style.font = Font(size=10, name="Calibri", color="1A1A2E")
    data_even_style.fill = PatternFill("solid", fgColor="FAFBFC")
    data_even_style.border = Border(
        left=Side(style="thin", color="D5DCE4"),
        right=Side(style="thin", color="D5DCE4"),
        top=Side(style="thin", color="D5DCE4"),
        bottom=Side(style="thin", color="D5DCE4")
    )
    styles['data_even'] = data_even_style
    
    # Ячейка данных (нечётная строка)
    data_odd_style = NamedStyle(name="data_odd_style")
    data_odd_style.font = Font(size=10, name="Calibri", color="1A1A2E")
    data_odd_style.fill = PatternFill("solid", fgColor="FFFFFF")
    data_odd_style.border = Border(
        left=Side(style="thin", color="D5DCE4"),
        right=Side(style="thin", color="D5DCE4"),
        top=Side(style="thin", color="D5DCE4"),
        bottom=Side(style="thin", color="D5DCE4")
    )
    styles['data_odd'] = data_odd_style
    
    # Итоговая строка
    total_style = NamedStyle(name="total_style")
    total_style.font = Font(bold=True, size=10, color="1A1A2E", name="Calibri")
    total_style.fill = PatternFill("solid", fgColor="EBF5FB")
    total_style.border = Border(
        left=Side(style="thin", color="D5DCE4"),
        right=Side(style="thin", color="D5DCE4"),
        top=Side(style="thin", color="D5DCE4"),
        bottom=Side(style="thin", color="D5DCE4")
    )
    styles['total'] = total_style
    
    # Карточка KPI
    kpi_label_style = NamedStyle(name="kpi_label_style")
    kpi_label_style.font = Font(size=9, color="888888", name="Calibri")
    kpi_label_style.alignment = Alignment(horizontal="center", vertical="center")
    kpi_label_style.fill = PatternFill("solid", fgColor="FFFFFF")
    styles['kpi_label'] = kpi_label_style
    
    kpi_value_style = NamedStyle(name="kpi_value_style")
    kpi_value_style.font = Font(bold=True, size=16, color="C0392B", name="Calibri")
    kpi_value_style.alignment = Alignment(horizontal="center", vertical="center")
    kpi_value_style.fill = PatternFill("solid", fgColor="FADBD8")
    styles['kpi_value'] = kpi_value_style
    
    return styles

def apply_cell_style(cell, style):
    """Применяет стиль к ячейке"""
    for attr in ['font', 'fill', 'alignment', 'border', 'number_format']:
        if hasattr(style, attr):
            setattr(cell, attr, getattr(style, attr))

# KPI КАРТОЧКИ
def add_kpi_row(ws, start_row, kpis, cols=6):
    """Добавляет строку с KPI карточками"""
    # Ряд с метками
    for i, (label, _) in enumerate(kpis):
        col = i * 2 + 1
        if col > cols:
            break
        ws.merge_cells(start_row=start_row, start_column=col, 
                       end_row=start_row, end_column=col+1)
        cell = ws.cell(row=start_row, column=col)
        cell.value = label
        apply_cell_style(cell, get_styles()['kpi_label'])
    
    # Ряд со значениями
    for i, (_, value) in enumerate(kpis):
        col = i * 2 + 1
        if col > cols:
            break
        ws.merge_cells(start_row=start_row+1, start_column=col,
                       end_row=start_row+1, end_column=col+1)
        cell = ws.cell(row=start_row+1, column=col)
        cell.value = value
        apply_cell_style(cell, get_styles()['kpi_value'])
    
    return start_row + 2

# ГРАФИКИ
def add_bar_chart(ws, data_row_start, data_row_end, col_categories, col_values, 
                  title, target_cell, chart_title):
    """Добавляет столбчатую диаграмму"""
    chart = BarChart()
    chart.title = chart_title
    chart.style = 10
    chart.width = 12
    chart.height = 8
    
    # Данные
    data = Reference(ws, min_col=col_values, min_row=data_row_start, 
                     max_row=data_row_end, max_col=col_values)
    cats = Reference(ws, min_col=col_categories, min_row=data_row_start+1,
                     max_row=data_row_end, max_col=col_categories)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Стилизация
    chart.x_axis.title = "Период"
    chart.y_axis.title = title
    chart.legend.position = 'b'
    
    ws.add_chart(chart, target_cell)

def add_line_chart(ws, data_row_start, data_row_end, col_categories, col_values,
                   title, target_cell, chart_title):
    """Добавляет линейную диаграмму"""
    chart = LineChart()
    chart.title = chart_title
    chart.style = 12
    chart.width = 12
    chart.height = 8
    
    data = Reference(ws, min_col=col_values, min_row=data_row_start,
                     max_row=data_row_end, max_col=col_values)
    cats = Reference(ws, min_col=col_categories, min_row=data_row_start+1,
                     max_row=data_row_end, max_col=col_categories)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend.position = 'b'
    
    ws.add_chart(chart, target_cell)

# ОТЧЁТ ПО КЛИЕНТАМ
def build_clients_report(wb, data, date_from, date_to):
    """Строит отчёт по клиентам (обновлённая версия)"""
    summary = data.get("summary", {})
    dynamics = data.get("dynamics", [])
    top_clients = data.get("topClients", [])
    
    styles = get_styles()
    
    # Основной лист
    ws = wb.active
    ws.title = "Клиенты"
    ws.sheet_view.showGridLines = False
    
    # НОВАЯ ШИРИНА КОЛОНОК
    column_widths = [8, 12, 14, 14, 12, 12, 18, 14, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ЗАГОЛОВОК
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.merge_cells('A2:I2')
    title_cell = ws['A2']
    title_cell.value = "ОТЧЁТ ПО КЛИЕНТАМ"
    apply_cell_style(title_cell, styles['title'])
    
    ws.row_dimensions[3].height = 20
    ws.merge_cells('A3:I3')
    period_cell = ws['A3']
    period_cell.value = f"Период: {format_date_range(date_from, date_to)}"
    apply_cell_style(period_cell, styles['period'])
    
    ws.row_dimensions[4].height = 10
    
    # KPI КАРТОЧКИ
    kpis = [
        ("Всего клиентов", format_number(summary.get("TotalClients", 0))),
        ("Новых клиентов", format_number(summary.get("NewClients", 0))),
        ("Всего заявок", format_number(summary.get("TotalApplications", 0))),
        ("Заказов", format_number(summary.get("TotalOrders", 0))),
        ("Общая Стоимость", format_currency(summary.get("TotalRevenue", 0))),
        ("Средний чек", format_currency(summary.get("AvgOrderValue", 0))),
    ]
    add_kpi_row(ws, 5, kpis, 8)
    ws.row_dimensions[7].height = 10
    
    # ДИНАМИКА С НАКОПЛЕНИЕМ
    if dynamics:
        ws.row_dimensions[8].height = 24
        ws.merge_cells('A8:I8')
        section_cell = ws['A8']
        section_cell.value = "ДИНАМИКА КЛИЕНТОВ ПО ПЕРИОДАМ"
        apply_cell_style(section_cell, styles['section'])
        
        # НОВЫЕ ЗАГОЛОВКИ
        headers = ["Год", "Месяц", "Новых клиентов", "Всего клиентов", "Активные клиенты", "Заявок", "Заказов", "Стоимость, BYN", "Ср. чек, BYN"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            apply_cell_style(cell, styles['header'])
        ws.row_dimensions[9].height = 28
        
        # Данные с накоплением
        row = 10
        cumulative_total = 0
        total_new = total_apps = total_orders = total_revenue = 0
        
        for i, item in enumerate(dynamics):
            month_ru = translate_month(item.get("MonthName", ""))
            year = item.get("Year", "")
            period_label = f"{month_ru} {year}" if month_ru else str(year)
            
            new_clients = item.get("NewClients", 0)
            cumulative_total += new_clients
            apps = item.get("Applications", 0)
            orders = item.get("Orders", 0)
            revenue = item.get("Revenue", 0)
            avg_check = revenue / cumulative_total if cumulative_total > 0 else 0
            active_clients = item.get("ActiveClients", cumulative_total)  # если нет данных, используем накопленные
            
            total_new += new_clients
            total_apps += apps
            total_orders += orders
            total_revenue += revenue
            
            cells = [year, period_label, new_clients, cumulative_total, active_clients, apps, orders, revenue, avg_check]
            style = styles['data_even'] if (row - 10) % 2 == 0 else styles['data_odd']
            
            for col, val in enumerate(cells, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = val
                apply_cell_style(cell, style)
                if col >= 3:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if col == 8 or col == 9:
                    cell.number_format = '#,##0.00'
            row += 1
        
        # Итоговая строка
        total_cell = ws.cell(row=row, column=1)
        total_cell.value = "ИТОГО"
        apply_cell_style(total_cell, styles['total'])
        
        for col, val in [(3, total_new), (4, cumulative_total), (6, total_apps), (7, total_orders)]:
            cell = ws.cell(row=row, column=col)
            cell.value = val
            apply_cell_style(cell, styles['total'])
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        cell = ws.cell(row=row, column=8)
        cell.value = total_revenue
        apply_cell_style(cell, styles['total'])
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '#,##0.00'
        
        avg_check_total = total_revenue / cumulative_total if cumulative_total > 0 else 0
        cell = ws.cell(row=row, column=9)
        cell.value = avg_check_total
        apply_cell_style(cell, styles['total'])
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '#,##0.00'
        
        ws.row_dimensions[row].height = 22
    
    # ТОП КЛИЕНТОВ
    if top_clients:
        ws2 = wb.create_sheet("Топ клиентов")
        ws2.sheet_view.showGridLines = False
        
        # НОВАЯ ШИРИНА КОЛОНОК
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 28
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 15
        ws2.column_dimensions['G'].width = 15
        
        ws2.row_dimensions[1].height = 10
        ws2.row_dimensions[2].height = 32
        ws2.merge_cells('A2:G2')
        title_cell = ws2['A2']
        title_cell.value = "ТОП КЛИЕНТОВ ПО СУММЕ ЗАКАЗОВ"
        apply_cell_style(title_cell, styles['title'])
        
        ws2.row_dimensions[3].height = 20
        ws2.merge_cells('A3:G3')
        period_cell = ws2['A3']
        period_cell.value = f"Период: {format_date_range(date_from, date_to)}"
        apply_cell_style(period_cell, styles['period'])
        
        # НОВЫЕ ЗАГОЛОВКИ (средний чек)
        headers = ["ФИО клиента", "Email", "Телефон", "Заказов", "Сумма, BYN", "Ср. чек, BYN", "Последний заказ"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=5, column=col)
            cell.value = header
            apply_cell_style(cell, styles['header'])
        ws2.row_dimensions[5].height = 28
        
        row = 6
        for i, client in enumerate(top_clients[:10]):
            style = styles['data_even'] if i % 2 == 0 else styles['data_odd']
            
            orders_count = client.get("OrdersCount", 0)
            total_spent = client.get("TotalSpent", 0)
            avg_check = total_spent / orders_count if orders_count > 0 else 0
            
            cells = [
                client.get("FullName", "—"),
                client.get("Email", "—"),
                client.get("Phone", "—"),
                orders_count,
                total_spent,
                avg_check,
                fmt_date(client.get("LastOrderDate")),
            ]
            
            for col, val in enumerate(cells, 1):
                cell = ws2.cell(row=row, column=col)
                cell.value = val
                apply_cell_style(cell, style)
                if col == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 5 or col == 6:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
        
        ws2.row_dimensions[row].height = 22
        total_cell = ws2.cell(row=row, column=1)
        total_cell.value = f"ИТОГО В ТОПЕ: {min(len(top_clients), 10)} КЛИЕНТОВ"
        apply_cell_style(total_cell, styles['total'])
        ws2.merge_cells(f'A{row}:G{row}')

#  ОТЧЁТ ПО ЗАКАЗАМ
def build_orders_report(wb, data, date_from, date_to):
    """Строит отчёт по заказам"""
    if isinstance(data, list):
        orders = data
        summary = {}
        dynamics = []
    else:
        orders = data.get("orders", [])
        summary = data.get("summary", {})
        dynamics = data.get("dynamics", [])
    
    styles = get_styles()
    
    ws = wb.active
    ws.title = "Заказы"
    ws.sheet_view.showGridLines = False
    
    column_widths = [8, 18, 16, 16, 18, 16, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ЗАГОЛОВОК
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.merge_cells('A2:G2')
    title_cell = ws['A2']
    title_cell.value = "ОТЧЁТ ПО ЗАКАЗАМ"
    apply_cell_style(title_cell, styles['title'])
    
    ws.row_dimensions[3].height = 20
    ws.merge_cells('A3:G3')
    period_cell = ws['A3']
    period_cell.value = f"Период: {format_date_range(date_from, date_to)}"
    apply_cell_style(period_cell, styles['period'])
    
    ws.row_dimensions[4].height = 10
    
    # KPI
    kpis = [
        ("Всего заказов", format_number(summary.get("TotalOrders", 0))),
        ("Завершено", format_number(summary.get("CompletedOrders", 0))),
        ("В работе", format_number(summary.get("InProgressOrders", 0))),
        ("Уникальных клиентов", format_number(summary.get("UniqueClients", 0))),
    ]
    add_kpi_row(ws, 5, kpis, 7)
    ws.row_dimensions[7].height = 10
    
    # ДИНАМИКА
    if dynamics:
        ws.row_dimensions[8].height = 24
        ws.merge_cells('A8:G8')
        section_cell = ws['A8']
        section_cell.value = "ДИНАМИКА ПО МЕСЯЦАМ"
        apply_cell_style(section_cell, styles['section'])
        
        headers = ["Год", "Месяц", "Заказов", "Завершено", "Клиентов"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            apply_cell_style(cell, styles['header'])
        ws.row_dimensions[9].height = 28
        
        row = 10
        total_orders = total_completed = 0
        
        for item in dynamics:
            month_ru = translate_month(item.get("MonthName", ""))
            period = f"{month_ru} {item.get('Year', '')}" if month_ru else str(item.get("Year", ""))
            
            orders_cnt = item.get("OrdersCount", 0)
            completed = item.get("CompletedCount", 0)
            clients = item.get("ClientsCount", 0)
            
            total_orders += orders_cnt
            total_completed += completed
            
            cells = [item.get("Year", ""), period, orders_cnt, completed, clients]
            style = styles['data_even'] if (row - 10) % 2 == 0 else styles['data_odd']
            
            for col, val in enumerate(cells, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = val
                apply_cell_style(cell, style)
                if col >= 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        
        total_cell = ws.cell(row=row, column=1)
        total_cell.value = "ИТОГО"
        apply_cell_style(total_cell, styles['total'])
        
        for col, val in [(3, total_orders), (4, total_completed)]:
            cell = ws.cell(row=row, column=col)
            cell.value = val
            apply_cell_style(cell, styles['total'])
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[row].height = 22
    
    # СПИСОК ЗАКАЗОВ
    if orders:
        ws2 = wb.create_sheet("Список заказов")
        ws2.sheet_view.showGridLines = False
        
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 28
        ws2.column_dimensions['D'].width = 28
        ws2.column_dimensions['E'].width = 22
        ws2.column_dimensions['F'].width = 14
        ws2.column_dimensions['G'].width = 20
        
        ws2.row_dimensions[1].height = 10
        ws2.row_dimensions[2].height = 32
        ws2.merge_cells('A2:G2')
        title_cell = ws2['A2']
        title_cell.value = "СПИСОК ЗАКАЗОВ"
        apply_cell_style(title_cell, styles['title'])
        
        ws2.row_dimensions[3].height = 20
        ws2.merge_cells('A3:G3')
        period_cell = ws2['A3']
        period_cell.value = f"Период: {format_date_range(date_from, date_to)}"
        apply_cell_style(period_cell, styles['period'])
        
        headers = ["№ Договора", "Дата", "Клиент", "Объект", "Специалист", "Статус", "Тип объекта"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=5, column=col)
            cell.value = header
            apply_cell_style(cell, styles['header'])
        ws2.row_dimensions[5].height = 28
        
        row = 6
        for i, order in enumerate(orders):
            status = order.get("Status", "")
            style = styles['data_even'] if i % 2 == 0 else styles['data_odd']
            
            cells = [
                order.get("ContractNumber", "—"),
                fmt_date(order.get("SignDate")),
                order.get("ClientName", "—"),
                order.get("ObjectName", "—"),
                order.get("SpecialistName", "—"),
                status,
                order.get("ObjectType", "—"),
            ]
            
            for col, val in enumerate(cells, 1):
                cell = ws2.cell(row=row, column=col)
                cell.value = val
                apply_cell_style(cell, style)
                if col == 2 or col == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Подсветка завершённых заказов
            if "Завершена" in status:
                for col in range(1, 8):
                    cell = ws2.cell(row=row, column=col)
                    cell.fill = PatternFill("solid", fgColor="E9F7EF")
            row += 1
        
        ws2.row_dimensions[row].height = 22
        total_cell = ws2.cell(row=row, column=1)
        total_cell.value = f"ИТОГО ЗАКАЗОВ: {len(orders)}"
        apply_cell_style(total_cell, styles['total'])
        ws2.merge_cells(f'A{row}:G{row}')

# ОФОРМЛЕНИЕ КНИГИ
def style_workbook(wb, report_title, date_from, date_to):
    """Применяет общее оформление ко всем листам"""
    date_str = format_date_range(date_from, date_to)
    
    for ws in wb.worksheets:
        # Колонтитулы
        ws.oddFooter.center.text = f"MCK-Reliable | {report_title} | {date_str} | Стр. &P из &N"
        ws.oddFooter.center.size = 9
        ws.oddFooter.center.color = "888888"
        
        # Закрепление области
        ws.freeze_panes = ws.cell(row=10, column=1)
        
        # Поля страницы
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

# ТОЧКА ВХОДА
def main():
    if len(sys.argv) < 6:
        print("Usage: generate_report.py <type> <json_data> <date_from> <date_to> <output_path>")
        sys.exit(1)
    
    report_type = sys.argv[1]
    data_path = sys.argv[2]
    date_from = sys.argv[3]
    date_to = sys.argv[4]
    output_path = sys.argv[5]
    
    try:
        with open(data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"JSON parse error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error reading file: {e}", file=sys.stderr)
        sys.exit(1)
    
    wb = Workbook()
    
    if report_type == "clients":
        build_clients_report(wb, data, date_from, date_to)
        title = "Отчёт по клиентам"
    elif report_type == "orders":
        build_orders_report(wb, data, date_from, date_to)
        title = "Отчёт по заказам"
    elif report_type == "applications":
        # Для заявок пока используем клиентский отчёт
        build_clients_report(wb, data, date_from, date_to)
        title = "Отчёт по заявкам"
    elif report_type == "specialists":
        # Для специалистов пока используем заказы
        build_orders_report(wb, data, date_from, date_to)
        title = "Отчёт по специалистам"
    else:
        print(f"Unknown report type: {report_type}", file=sys.stderr)
        sys.exit(1)
    
    style_workbook(wb, title, date_from, date_to)
    
    wb.save(output_path)
    print(f"Saved: {output_path}")

if __name__ == "__main__":
    main()